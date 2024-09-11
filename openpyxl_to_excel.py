
# Loop Through Recent Articles for Article Link, Author Name, Author Page

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
import time
import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import xlsxwriter

# Set up a Selenium WebDriver instance
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# List of page URLs
url_defi = "https://thedefiant.io/news/DeFi"
url_cefi = "https://thedefiant.io/news/cefi"
url_tradfi_fintech = "https://thedefiant.io/news/tradfi-and-fintech"
url_blockchains = "https://thedefiant.io/news/blockchains"
url_nfts_web3 = "https://thedefiant.io/news/nfts-and-web3"
url_people = "https://thedefiant.io/news/people"
url_markets = "https://thedefiant.io/news/markets"
url_regulation = "https://thedefiant.io/news/regulation"
url_hacks = "https://thedefiant.io/news/hacks"
url_research_opinion = "https://thedefiant.io/news/research-and-opinion"
url_pressreleases = "https://thedefiant.io/news/press-releases"
url_sponsored = "https://thedefiant.io/news/sponsored"
url_deepnewz = "https://thedefiant.io/news/deep-newz"

# As a list to loop through
url_list = [url_defi, url_cefi, url_tradfi_fintech, url_blockchains, url_nfts_web3, url_people,
url_markets, url_regulation, url_hacks, url_research_opinion, url_pressreleases, url_sponsored, url_deepnewz ]


# wb = openpyxl.load_workbook("test1.xlsx") # Load workbook
# sheet_1 = wb.active

wb = Workbook()
ws0 = wb.worksheets[0]
ws0.title = 'My Sheet 1'
ws1 = wb.create_sheet()
ws1.title = 'My Sheet 2'
ws2 = wb.create_sheet()
ws2.title = 'My Sheet 3'


# Loop through pages, links, authors, author links
for url in url_list:
    driver.get(url)

    # Wait for the page to load 
    time.sleep(1) 

    # Get list of Recent Articles as web elements
    links_list = [link for link in driver.find_elements(By.XPATH, "//section[@class='mt-4']/div//div[@class='flex flex-col']/a")]

    # Get list of Recent Articles as links
    recent_article_links = []
    for link in links_list:
        attr_link = link.get_attribute("href")
        recent_article_links.append(attr_link)
    print(recent_article_links)

# Loop through links
    author_links = []
    author_names = []
    for link in recent_article_links:
        driver.get(link)
        author_link = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, "//article/div[2]/a[@href]"))).get_attribute("href")  # Get author link
        author_name = driver.find_element(By.XPATH, "//article/div[2]/a[@href]").text
        author_links.append(author_link)
        author_names.append(author_name)

    # Get unique author links (set) for each loop
    unique_author_set = set(author_links)
    unique_author_set_list = list(unique_author_set)
    ws1.append(unique_author_set_list)   #! Ideally this will append to worksheed index 1 (My Sheet 2)
    
    #! PROBABLY ISSUE HERE
    #TODO: Click on author link 

    # recent_article_links = []
    # for link in links_list:
    #     attr_link = link.get_attribute("href")
    #     recent_article_links.append(attr_link)


    author_article_links = []
    author_article_text = []
    for author_link in unique_author_set_list:
        driver.get(author_link)
        author_articles = driver.find_elements(By.XPATH, '/html/body/div[1]/main/div[2]//div/a[@href]') #TODO: Can cut off @href here, add get_attribute in a for loop
        for article_title in author_articles:
            links = article_title.get_attribute("href") #! Error might be here
            article_titles = article_title.text
            #TODO: Add Href get attribute here
            author_article_text.append(article_titles)
            author_article_links.append(links)
        print('Author articles are', author_article_links)
        print('Author article texts are', author_article_text)
        #TODO: Have text here, but need links as well?
        #TODO: Output both to excel ws2 (My Sheet 3)

    ws2.append(author_article_links)
    ws2.append(author_article_text)

    #? FULL XPATH for Author Articles: /html/body/div[1]/main/div[2]/div/div[1]/div/a


    nested_list = [recent_article_links, author_names, author_links]

    # Append sublists to separate rows
    for sub_list in nested_list:
        ws0.append(sub_list)

# # Append row category (Defi, Cefi, etc.)
# col1 = sheet_1.insert_cols(0, amount=1)
# sheet_1.append(url_list)

#TODO: Adjust sheet to add Title before each excel row (DeFi, CeFi, Tradfi, NFTs, etc.)

#! Need to solve

#TODO: Set of author names (to kick out repeats, and click on once for each)



#TODO: Then scrape list of articles by author

wb.save("test1.xlsx")    # Save workbook




